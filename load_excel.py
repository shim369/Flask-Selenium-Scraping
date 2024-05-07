import openpyxl

def load_excel(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    max_row = ws.max_row + 1
    data = []
    for i in range(2,max_row):
        index = ws.cell(row=i,column=1).value
        name = ws.cell(row=i,column=2).value
        gender = ws.cell(row=i,column=4).value
        pref = ws.cell(row=i,column=8).value
        blood = ws.cell(row=i,column=14).value
        data.append([index,name,gender,pref,blood])
    return data

if __name__ == "__main__":
    print(load_excel("persons.xlsx"))